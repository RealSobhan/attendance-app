import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side
import os
import requests
import time
import streamlit as st # Import Streamlit

#======================================================================
# FUNCTION 1: GETS DATA FROM THE API
# This is the same function from the previous version.
#======================================================================
def get_calendar_from_api(year, month):
    """
    Gets calendar data for a specific Persian month and year using the holidayapi.ir API.
    This version has NO external date library dependencies. It finds the first available
    weekday from the API in the first week and calculates the rest.
    """
    progress_bar = st.progress(0, text="در حال دریافت اطلاعات از سرور تقویم...")
    
    persian_months = {
        1: 'فروردین', 2: 'اردیبهشت', 3: 'خرداد',
        4: 'تیر', 5: 'مرداد', 6: 'شهریور',
        7: 'مهر', 8: 'آبان', 9: 'آذر',
        10: 'دی', 11: 'بهمن', 12: 'اسفند'
    }
    month_name = persian_months.get(month, '')
    if not month_name:
        st.error(f"ماه نامعتبر است: {month}. لطفا عددی بین 1 تا 12 وارد کنید.")
        return None

    weekdays = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه']
    valid_weekdays = set(weekdays)
    start_weekday_index = -1
    anchor_day = -1

    for day_to_check in range(1, 8):
        try:
            url_check = f'https://holidayapi.ir/jalali/{year}/{month}/{day_to_check}'
            response_check = requests.get(url_check)
            if response_check.status_code in [404, 400]:
                break
            response_check.raise_for_status()
            data_check = response_check.json()
            if data_check.get('events'):
                for event in data_check['events']:
                    if event.get('description') in valid_weekdays:
                        start_weekday_index = weekdays.index(event['description'])
                        anchor_day = day_to_check
                        break
            if start_weekday_index != -1:
                break
        except requests.exceptions.RequestException as e:
            st.error(f"خطا در ارتباط با سرور تقویم: {e}")
            return None
        except ValueError:
            continue

    if start_weekday_index == -1:
        st.error(f"خطا: روز شروع هفته برای ماه {month} سال {year} یافت نشد.")
        return None

    school_days = []
    for day in range(1, 32):
        url = f'https://holidayapi.ir/jalali/{year}/{month}/{day}'
        try:
            response = requests.get(url)
            if response.status_code in [404, 400]:
                break
            response.raise_for_status()
            data = response.json()
            
            is_holiday = data.get('is_holiday', False)
            current_weekday_index = (start_weekday_index + (day - anchor_day)) % 7
            weekday_name = weekdays[current_weekday_index]

            if not is_holiday and weekday_name not in ['پنجشنبه', 'جمعه']:
                formatted_date = f"{year}/{str(month).zfill(2)}/{str(day).zfill(2)}"
                school_days.append({
                    'persian_month': month_name,
                    'persian_weekday': weekday_name,
                    'formatted_date': formatted_date
                })
            time.sleep(0.05)
            progress_bar.progress(day / 31, text=f"در حال بررسی روز {day}...")


        except requests.exceptions.RequestException:
            continue
        except ValueError:
            continue
            
    if not school_days:
        st.warning(f"هیچ روز درسی برای ماه {month} سال {year} یافت نشد.")
        return None

    progress_bar.empty()
    st.write("اطلاعات با موفقیت دریافت شد.")
    return pd.DataFrame(school_days)

#======================================================================
# FUNCTION 2: GENERATES THE EXCEL FILE
# This is the same function from the previous version.
#======================================================================
def generate_minimal_attendance_sheet(dates_df, 
                                      font_sizes,
                                      column_widths,
                                      row_height,
                                      filename):
    """
    Generates a single Excel sheet with a minimal attendance log from a pandas DataFrame.
    """
    if dates_df is None or dates_df.empty:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Log"
    ws.sheet_view.rightToLeft = True
    
    main_header_font = Font(name='B Zar', size=font_sizes.get('main_header', 16), bold=True)
    header_font = Font(name='B Zar', size=font_sizes.get('header', 12), bold=True)
    cell_font = Font(name='B Zar', size=font_sizes.get('cell', 11))
    date_font = Font(name='B Zar', size=font_sizes.get('date', 10))
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border_side = Side(style='thin')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    month_name = dates_df['persian_month'].iloc[0]
    ws.merge_cells('A1:F1')
    month_header_cell = ws['A1']
    month_header_cell.value = month_name
    month_header_cell.font = main_header_font
    month_header_cell.alignment = center_alignment
    month_header_cell.border = thin_border
    # Set a fixed height for the main header
    ws.row_dimensions[1].height = 45 

    table_start_row = 2
    headers = ["روز هفته", "تاریخ", "زنگ", "اسامی غایبین", "اسامی و میزان تاخیر", "نام و امضای دبیر"]
    for col, header_text in enumerate(headers, 1):
        cell = ws.cell(row=table_start_row, column=col, value=header_text)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    # Set a fixed height for the column headers
    ws.row_dimensions[table_start_row].height = 35

    current_row = table_start_row + 1
    
    for index, row_data in dates_df.iterrows():
        weekday = row_data['persian_weekday']
        
        if weekday in ['دوشنبه', 'چهارشنبه']:
            num_periods = 3
            periods = ["اول", "دوم", "سوم"]
        else:
            num_periods = 4
            periods = ["اول", "دوم", "سوم", "چهارم"]

        day_merge_start = current_row
        day_merge_end = current_row + num_periods - 1

        ws.merge_cells(start_row=day_merge_start, start_column=1, end_row=day_merge_end, end_column=1)
        ws.merge_cells(start_row=day_merge_start, start_column=2, end_row=day_merge_end, end_column=2)

        weekday_cell = ws.cell(row=day_merge_start, column=1, value=weekday)
        weekday_cell.font = date_font
        weekday_cell.alignment = center_alignment

        date_cell = ws.cell(row=day_merge_start, column=2, value=row_data['formatted_date'])
        date_cell.font = date_font
        date_cell.alignment = center_alignment
        
        for i in range(num_periods):
            # Apply the user-selected height only to data rows
            ws.row_dimensions[current_row + i].height = row_height
            period_cell = ws.cell(row=current_row + i, column=3, value=periods[i])
            period_cell.font = cell_font
            period_cell.alignment = center_alignment
            for col_num in range(1, 7):
                ws.cell(row=current_row + i, column=col_num).border = thin_border
        
        current_row += num_periods

    # --- Hide unused rows and columns for a cleaner look ---
    ws.sheet_view.showGridLines = False
    # Set print area to only the generated table
    ws.print_area = f'A1:F{current_row - 1}'
    # Hide all columns after 'F'
    for i in range(7, 200):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].hidden = True
    # Hide all rows after the last data row
    for i in range(current_row, 500):
        ws.row_dimensions[i].hidden = True


    # Save the workbook to a temporary file in memory
    from io import BytesIO
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook

#======================================================================
# STREAMLIT USER INTERFACE (REDESIGNED)
#======================================================================

# Use wide layout for the page
st.set_page_config(layout="wide", page_title="سازنده فرم حضور و غیاب", page_icon="📄")

# --- CUSTOM FONT, DARK THEME, AND RTL CSS ---
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Vazirmatn:wght@400;700&display=swap');
    
    html, body, [class*="st-"], [class*="css-"] {
        font-family: 'Vazirmatn', sans-serif;
        direction: rtl;
        font-size: 1.1rem; /* Increase base font size */
    }
    
    /* Main app background */
    [data-testid="stAppViewContainer"] > .main {
        background-color: #0E1117;
        color: #FAFAFA;
        padding: 2rem 3rem;
    }

    /* Sidebar background and spacing */
    [data-testid="stSidebar"] {
        background-color: #1a1c24;
        padding: 1.5rem;
        width: 400px !important; /* Increase sidebar width */
    }
    
    /* Headers */
    h1 {
        font-family: 'Vazirmatn', sans-serif;
        font-weight: 700;
        font-size: 2.8rem; /* Increase title size */
        text-shadow: 2px 2px 8px rgba(0,0,0,0.3);
    }
    h2 {
        font-size: 2.2rem;
    }
    h3 {
        font-size: 1.6rem;
        font-weight: 700;
    }
    
    /* Make buttons more visible */
    .stButton > button {
        border: 2px solid #00b894;
        background-color: #00b894;
        color: white;
        font-weight: 700;
        padding: 0.75rem 1rem;
        font-size: 1.2rem;
        border-radius: 10px;
    }
    .stButton > button:hover {
        border: 2px solid #00a884;
        background-color: #00a884;
        color: white;
    }
    
    /* Style for the results container */
    .st-emotion-cache-1jicfl2 {
        background-color: #1a1c24;
        border: 1px solid #444;
        border-radius: 10px;
        padding: 1.5rem;
    }
    </style>
    """, unsafe_allow_html=True)

# --- SIDEBAR CONTROLS (ORGANIZED) ---
with st.sidebar:
    st.header("🗓️ تنظیمات اصلی")
    target_year = st.number_input("سال (مثلا: 1404)", min_value=1390, max_value=1450, value=1404)
    target_month = st.number_input("ماه (مثلا: برای مهر 7)", min_value=1, max_value=12, value=7)
    
    st.markdown("---")

    st.header("🎨 تنظیمات ظاهری")
    header_font_size = st.slider("اندازه فونت سربرگ", 10, 24, 14, key="hfs")
    cell_font_size = st.slider("اندازه فونت داخلی", 10, 24, 12, key="cfs")
    col_d_width = st.slider("عرض ستون 'غایبین'", 30, 80, 50, key="cdw")
    custom_row_height = st.slider("ارتفاع ردیف (فاصله عمودی)", 20, 50, 30, key="crh")
    
    st.markdown("---")
    
    generate_button = st.button("🚀 ساخت فایل اکسل", type="primary", use_container_width=True)


# --- MAIN PAGE ---
st.title("📄 سازنده فرم حضور و غیاب")
st.markdown("این برنامه به شما کمک می‌کند تا به سرعت فرم حضور و غیاب ماهانه برای کلاس خود بسازید.")

# --- NEW INFORMATION SECTION (UPDATED) ---
with st.expander("ℹ️ راهنمای تبدیل فایل اکسل به PDF"):
    st.write("""
        برای دریافت خروجی PDF با بهترین کیفیت از فایل اکسل، مراحل زیر را دنبال کنید:

        **مرحله ۱: تعیین محدوده پرینت**
        - پس از باز کردن فایل اکسل، به تب **Page Layout** بروید.
        - با موس، کل جدول را از اولین تا آخرین سلول انتخاب کنید.
        - روی گزینه **Print Area** کلیک کرده و سپس **Set Print Area** را انتخاب نمایید.

        **مرحله ۲: تنظیمات صفحه**
        - در همان تب **Page Layout**، در بخش "Scale to Fit"، مقدار **Width** را روی **1 page** و مقدار **Height** را روی **2 pages** تنظیم کنید. این کار باعث می‌شود تمام ستون‌ها در یک صفحه جا شوند.

        **مرحله ۳: تکرار سربرگ در صفحات**
        - روی گزینه **Print Titles** کلیک کنید.
        - در پنجره باز شده، روی فلش کنار کادر **Rows to repeat at top** کلیک کنید.
        - سطر حاوی عنوان ستون‌ها (مثلا: ردیف ۲) را انتخاب کرده و Enter را بزنید.

        **مرحله ۴: اضافه کردن شماره صفحه**
        - در همان پنجره **Print Titles**، به تب **Header/Footer** بروید.
        - روی **Custom Footer** کلیک کرده و در بخش مورد نظر (مثلا: Center section)، شماره صفحه را اضافه کنید.

        با انجام این تنظیمات، هنگام پرینت گرفتن یا ذخیره به صورت PDF، فایل شما ظاهری مرتب و حرفه‌ای خواهد داشت.
    """)

st.markdown("---")

# Use columns for a cleaner layout - UPDATED RATIO
col1, col2 = st.columns([3, 2])

with col1:
    st.subheader("راهنما")
    st.write("""
    1.  از منوی کنار صفحه، **سال** و **ماه** مورد نظر خود را وارد کنید.
    2.  در صورت تمایل، **تنظیمات ظاهری** را تغییر دهید.
    3.  روی دکمه **ساخت فایل اکسل** کلیک کنید.
    4.  منتظر بمانید تا فایل ساخته شود و سپس دکمه **دانلود** در قسمت نتایج ظاهر خواهد شد.
    """)

with col2:
    if generate_button:
        # This is where the results will appear, only after the button is clicked
        with st.container(border=True):
            st.subheader("نتایج")
            # 1. Get the data from the API
            dates_dataframe = get_calendar_from_api(target_year, target_month)

            # 2. Generate the Excel file if data was found
            if dates_dataframe is not None and not dates_dataframe.empty:
                
                # Define styles based on advanced options
                custom_font_sizes = {'header': header_font_size, 'cell': cell_font_size, 'date': cell_font_size, 'main_header': 18}
                custom_column_widths = {'A': 15, 'B': 12, 'C': 8, 'D': col_d_width, 'E': col_d_width, 'F': 25}
                
                month_name_for_file = dates_dataframe['persian_month'].iloc[0]
                output_filename = f"فرم_حضور_غیاب_{month_name_for_file}_{target_year}.xlsx"

                # Generate the file in memory
                excel_data = generate_minimal_attendance_sheet(
                    dates_dataframe,
                    font_sizes=custom_font_sizes,
                    column_widths=custom_column_widths,
                    row_height=custom_row_height,
                    filename=output_filename
                )
                
                if excel_data:
                    st.success(f"✅ فایل '{output_filename}' با موفقیت ساخته شد!")
                    
                    # Provide a download button
                    st.download_button(
                        label="📥 دانلود فایل اکسل",
                        data=excel_data,
                        file_name=output_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            else:
                st.error("خطا: اطلاعاتی برای این ماه دریافت نشد. لطفا سال و ماه دیگری را امتحان کنید.")
